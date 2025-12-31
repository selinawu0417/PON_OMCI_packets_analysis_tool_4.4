import tkinter as tk
from tkinter import Menu, ttk, scrolledtext, filedialog, messagebox, Toplevel
import os
import csv
import re
from tkinter.filedialog import asksaveasfilename, askopenfilename
import pandas as pd
from graphviz import Digraph
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles import Border, Side


class OMCIAnalyzerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("PON OMCI packets analysis tool v4.0")
        self.root.geometry("800x600")

        # Create Menu
        self.menu = Menu(root)
        root.config(menu=self.menu)

        # OMCI Menu
        self.omci_menu = Menu(self.menu, tearoff=0)
        self.menu.add_cascade(label="OMCI", menu=self.omci_menu)
        self.omci_menu.add_command(label="Open Wireshark File", command=self.open_wireshark_file)
        self.omci_menu.add_command(label="Generate OMCI Analysis Table", command=self.generate_omci_analysis_table)
      #  self.omci_menu.add_command(label="Generate Layer 2 OMCI Table", command=self.generate_layer2_omci_table)

        # Test Cases Menu
        self.test_cases_menu = Menu(self.menu, tearoff=0)
        self.menu.add_cascade(label="Test Cases", menu=self.test_cases_menu)
        self.test_cases_menu.add_command(label="Open TestRail CSV File", command=self.open_testrail_csv)


        self.help_menu = Menu(self.menu, tearoff=0)
        self.menu.add_cascade(label="Help", menu=self.help_menu)
        self.help_menu.add_command(label="Release note", command=self.open_release_note)


        # Search Bar and Buttons
        search_frame = tk.Frame(root)
        search_frame.pack(fill=tk.X, padx=10, pady=5)

        self.search_var = tk.StringVar()
        self.search_entry = tk.Entry(
            search_frame, textvariable=self.search_var, width=40, font=("Arial", 12)
        )
        self.search_entry.insert(0, "Please enter search keywords...")
        self.search_entry.bind("<FocusIn>", lambda e: self.search_entry.delete(0, tk.END))
        self.search_entry.pack(side=tk.LEFT, padx=5)

        search_button = tk.Button(search_frame, text="Search", command=self.search_text)
        search_button.pack(side=tk.LEFT, padx=5)

        prev_button = tk.Button(search_frame, text="Prev", command=self.search_prev)
        prev_button.pack(side=tk.LEFT, padx=5)

        next_button = tk.Button(search_frame, text="Next", command=self.search_next)
        next_button.pack(side=tk.LEFT, padx=5)

        self.search_count_label = tk.Label(search_frame, text="0 / 0", font=("Arial", 12))
        self.search_count_label.pack(side=tk.LEFT, padx=5)

        # Main Content Area (Dark Theme)
        self.packet_display = scrolledtext.ScrolledText(
            root, wrap=tk.WORD, width=80, height=20, bg="black", fg="white", insertbackground="white"
        )
        self.packet_display.pack(expand=True, fill=tk.BOTH, padx=10, pady=10)

        # Search tracking
        self.search_results = []
        self.current_search_index = -1

        # OMCI Data
        self.omci_packets = []

    def open_wireshark_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("Wireshark Files", "*.pcapng;*.pcap")])
        if not file_path:
            return
        self.omci_packets = self.filter_omci_packets(file_path)
        if self.omci_packets:
            self.packet_display.delete("1.0", tk.END)
            for packet in self.omci_packets:
                self.packet_display.insert(tk.END, f"{packet}\n")
            messagebox.showinfo("Info", "OMCI packets loaded successfully.")
        else:
            messagebox.showinfo("Info", "No OMCI packets found in the selected file.")

    def filter_omci_packets(self, file_path):
        try:
            import pyshark  # Import here to avoid errors in environments without pyshark
            capture = pyshark.FileCapture(file_path, display_filter="omci")
            omci_data = []
            for packet in capture:
                omci_data.append(str(packet))
            capture.close()
            return omci_data
        except Exception as e:
            messagebox.showerror("Error", f"Failed to filter OMCI packets: {str(e)}")
            return []

    def extract_omci_parameters(self):
        """
        Extract specified OMCI parameters from packets, including rules for ME Class Instance and ME Class Upload Content,
        ensuring all numeric values (including hexadecimal) are correctly displayed.
        """
        data = []
        for packet in self.omci_packets:
            # Extract fields using regex
            message_type = re.search(r"Message Type:\s*([\w\s]+)", packet)
            device_id = re.search(r"Device Identifier:\s*(0x\w+)", packet)
            #me_class = re.search(r"Message Identifier, ME Class =\s*([\w\s]+), Instance = (\d+)", packet)
            me_class = re.search(r"Message Identifier, ME Class =\s*([\w\s\-\.]+), Instance = (\d+)", packet)

            managed_entity_classes = re.findall(
               # r"Managed Entity Class:\s*([\w\s\-]+)\s*\(([\w\s]+)\)", packet
                r"Managed Entity Class:\s*([\w\s\-\.]+)\s*\(([\w\s]+)\)", packet
                
            )
            managed_entity_instances = re.findall(
                r"Managed Entity Instance:\s*(\d+|0x[\w]+)", packet
            )

            # Initialize fields
            me_class_upload_content = ""
            managed_entity_class_output = ""
            me_class_instance_content = ""

            # Handle ME Class Upload Content
            me_class_upload_start = re.search(r"ME Class Upload Content", packet)
            if me_class_upload_start:
                post_upload_content = packet[me_class_upload_start.end():]
                me_class_matches = re.findall(
                    r"Managed Entity Class:\s*([\w\s\-]+)\s*\(([\w\s]+)\)", post_upload_content
                )
                me_instance_matches = re.findall(
                    r"Managed Entity Instance:\s*(\d+|0x[\w]+)", post_upload_content
                )

                if me_class_matches:
                    me_class_upload_content = "ME Class Upload Content\n" + "\n".join(
                        f"Managed Entity Class: {cls} ({val})" for cls, val in me_class_matches
                    )

                if me_instance_matches:
                    # Include both decimal and hexadecimal representations
                    me_class_upload_content += "\n" + "\n".join(
                        f"Managed Entity Instance: {int(instance, 0)} (0x{int(instance, 0):04x})"
                        for instance in me_instance_matches
                    )

            # Handle Managed Entity Class field
            if managed_entity_classes:
                managed_entity_class_output = ", ".join(
                    f"{cls} ({val})" for cls, val in managed_entity_classes
                )

            # Handle ME Class Instance rules
            if managed_entity_instances or managed_entity_classes:
                # Add Managed Entity Instances
                if managed_entity_instances:
                    me_class_instance_content += "\n".join(
                        f"Managed Entity Instance: {int(instance, 0)} (0x{int(instance, 0):04x})"
                        for instance in managed_entity_instances
                    )
                # Add ME Class and Managed Entity Class
                if me_class:
                    me_class_instance_content += f"\nMessage Identifier, ME Class = {me_class.group(1)}, Instance = {me_class.group(2)}"
                if managed_entity_classes:
                    me_class_instance_content += "\n" + ", ".join(
                        f"Managed Entity Class: {cls} ({val})" for cls, val in managed_entity_classes
                    )

            # Extract Attribute Mask and all following content
            attribute_mask_match = re.search(r"Attribute Mask\s*\((0x\w+)\)", packet)
            attribute_mask_full = ""
            if attribute_mask_match:
                # Capture content inside parentheses
                attribute_mask_paren = f"({attribute_mask_match.group(1)})"
                # Capture multiline binary data following the Attribute Mask line
                start_pos = attribute_mask_match.start()
                mask_section = packet[start_pos:].splitlines()
                attribute_mask_full = f"{attribute_mask_paren}\n" + "\n".join(mask_section[1:]).strip()

            # Extract Attribute List and multiline content
            attribute_list_match = re.search(r"Attribute List\s*([\s\S]+?)(?=\n[A-Z]|$)", packet)
            attribute_list = ""
            if attribute_list_match:
                attribute_list_lines = attribute_list_match.group(1).strip().split("\n")
                attribute_list = "\n".join(line.strip() for line in attribute_list_lines)

            # Append extracted data
            data.append({
                "Message Type": message_type.group(1) if message_type else "",
                "Managed Entity Class": managed_entity_class_output.strip(),
                "ME Class Instance": me_class_instance_content.strip(),
                "ME Class Upload Content": me_class_upload_content.strip(),
                "Attribute Mask": attribute_mask_full,
                "Attribute List": attribute_list,
                "Device Identifier": device_id.group(1) if device_id else "",
            })
        return data









    def generate_omci_analysis_table(self):
        """
        Generate a well-formatted Excel table of OMCI data and automatically save to the root directory.
        """
        if not self.omci_packets:
            messagebox.showinfo("Info", "No OMCI packets loaded. Please open a Wireshark file first.")
            return

        # Extract data
        data = self.extract_omci_parameters()
        if not data:
            messagebox.showinfo("Info", "No parameters found in OMCI packets.")
            return

        try:
            # Convert data to a DataFrame
            df = pd.DataFrame(data)

            # Save file as OMCI_Analysis_Table.xlsx with option for save-as
            default_file_path = os.path.join(os.getcwd(), "OMCI_Analysis_Table.xlsx")
            save_path = asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel Files", "*.xlsx")],
                initialfile="OMCI_Analysis_Table.xlsx",
                initialdir=os.getcwd(),
            ) or default_file_path

            df.to_excel(save_path, index=False)

            # Apply formatting to the Excel file
            self.format_excel(save_path)

            messagebox.showinfo("Success", f"OMCI analysis table saved as {save_path}.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save OMCI analysis table: {str(e)}")


    def format_excel(self, file_path):
        """
        Apply formatting to the generated Excel file, including text alignment, color styling, and text wrapping.
        """
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Alignment, Font, PatternFill
            import re
            
            def hex_to_ascii(hex_str):
                """Convert hexadecimal string to ASCII, ignoring non-printable characters."""
                try:
                    bytes_object = bytes.fromhex(hex_str)
                    return bytes_object.decode("ascii", errors="ignore").strip()
                except Exception:
                    return ""            



            wb = load_workbook(file_path)

            ws = wb.active
            # Set default view to 80%
            ws.sheet_view.zoomScale = 80

            # Apply header formatting
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            border_style = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border_style



            # Apply formatting to data cells
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    # Set alignment to top-left
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                    # Apply border
                    cell.border = border_style

            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                col_letter = col[0].column_letter
                ws.column_dimensions[col_letter].width = min(max_length + 2, 50)  # Limit width for readability




            # List of parameters to highlight
            highlight_params = [
                "04: Expected Equipment Id",
                "03: Serial Number",
                "04: Version",
                "11: Total T-CONT Buffer Number",
                "12: Total Priority Queue Number",
                "13: Total Traffic Scheduler Number",
                "01: Version",
                "02: Is committed",
                "03: Is active",
                "04: Is valid",
                "01: OLT vendor id",
                "03: OLT version",
                "02: OMCC version"
            ]

        # Create a new sheet for highlighted content
            highlighted_ws = wb.create_sheet(title="Highlighted Content")
            highlighted_ws["A1"] = "Highlighted Content"
            highlighted_ws["A1"].font = Font(bold=True, color="FFFFFF")
            highlighted_ws["A1"].fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")

            # Apply formatting to data cells and copy highlighted content to the new worksheet
            highlighted_row = 2
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    # Highlight specific parameters
                    if any(param in str(cell.value) for param in highlight_params):
                        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow background
                        cell.font = Font(color="FF0000")  # Red text
                        

                        # Copy the highlighted content to the new sheet
                        highlighted_ws[f"A{highlighted_row}"] = cell.value
                        highlighted_ws[f"A{highlighted_row}"].font = Font(color="FF0000")
                        highlighted_ws[f"A{highlighted_row}"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                        highlighted_ws[f"A{highlighted_row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                        highlighted_row += 1

                    # Set alignment to top-left for original sheet
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)




            # Apply formatting to data cells
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    # Highlight specific parameters
                    if any(param in str(cell.value) for param in highlight_params):
                        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow background
                        cell.font = Font(color="FF0000")  # Red text

                    # Set alignment to top-left
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)




            # Apply formatting to data cells
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    # Set alignment to top-left
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

                    # Apply specific formatting based on column headers
                    header_value = ws.cell(row=1, column=cell.column).value
                    if header_value in ["Message Type", "Managed Entity Class","ME Class Instance"]:
                        cell.font = Font(color="0000FF")  # Blue
                    elif header_value in ["ME Class Upload Content", "Attribute Mask", "Attribute List"]:
                        cell.font = Font(color="A52A2A")  # Brown-red


            # Apply formatting to Highlighted Content sheet
            if "Highlighted Content" in wb.sheetnames:
                highlighted_ws = wb["Highlighted Content"]

                # Apply header formatting
                for cell in highlighted_ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.border = border_style

                # Apply formatting to data cells
                for row in highlighted_ws.iter_rows(min_row=2, max_row=highlighted_ws.max_row, max_col=highlighted_ws.max_column):
                    for cell in row:
                        # Set alignment to top-left
                        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                        # Apply border
                        cell.border = border_style

                # Auto-adjust column widths for the Highlighted Content sheet
                for col in highlighted_ws.columns:
                    max_length = 0
                    for cell in col:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    col_letter = col[0].column_letter
                    highlighted_ws.column_dimensions[col_letter].width = min(max_length + 2, 50)




            # Enable filters
            ws.auto_filter.ref = ws.dimensions

            # Save the formatted workbook
            wb.save(file_path)
         #   messagebox.showinfo("Success", f"Formatted Excel file with highlighted content saved as {file_path}.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to format Excel file: {str(e)}")


    def open_testrail_csv(self):
        file_path = filedialog.askopenfilename(filetypes=[("CSV Files", "*.csv")])
        if not file_path:
            return
        try:
            csv_data = {}
            with open(file_path, mode='r', encoding='utf-8') as file:
                reader = csv.DictReader(file)
                for row in reader:
                    csv_data[row['Title']] = row['Steps']

            self.display_test_cases(csv_data)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load CSV: {str(e)}")

    def display_test_cases(self, csv_data):
        test_case_window = Toplevel(self.root)
        test_case_window.title("Test Cases")
        test_case_window.geometry("600x400")

        label = tk.Label(test_case_window, text="Select a Test Case:", font=("Arial", 13))
        label.pack(pady=10)

        combo = ttk.Combobox(test_case_window, values=list(csv_data.keys()), state="readonly", font=("Arial", 12))
        combo.pack(pady=10, padx=20, fill=tk.X)

        def on_test():
            selected = combo.get()
            if selected:
                self.validate_test_case(csv_data[selected])

        test_button = tk.Button(test_case_window, text="Test", font=("Arial", 12), command=on_test)
        test_button.pack(pady=20)

    def validate_test_case(self, steps):
        steps_window = Toplevel(self.root)
        steps_window.title("Test Case Validation")
        steps_window.geometry("750x550")

        steps_text = scrolledtext.ScrolledText(steps_window, wrap=tk.WORD, width=60, height=10, font=("Arial", 12))
        steps_text.pack(expand=True, fill=tk.BOTH, padx=10, pady=10)
        steps_text.insert(tk.END, steps)
        steps_text.config(state=tk.DISABLED)

        result_label = tk.Label(steps_window, text="", font=("Arial", 14, "bold"))
        result_label.pack(pady=10)

        def validate_steps():
            steps_keywords = re.findall(r'"(.*?)"', steps, re.IGNORECASE)
            matches = []
            missing = []

            for keyword in steps_keywords:
                if any(keyword.lower() in packet.lower() for packet in self.omci_packets):
                    matches.append(keyword)
                else:
                    missing.append(keyword)

            if missing:
                result_label.config(text="FAIL", fg="red")
            else:
                result_label.config(text="PASS", fg="green")

            match_results_text.delete("1.0", tk.END)
            match_results_text.insert(tk.END, "\n".join(matches))

            missing_results_text.delete("1.0", tk.END)
            missing_results_text.insert(tk.END, "\n".join(missing))



        result_frame = tk.Frame(steps_window)
        result_frame.pack(expand=True, fill=tk.BOTH, padx=10, pady=10)

        validate_button = tk.Button(steps_window, text="Validate", font=("Arial", 12), command=validate_steps)
        validate_button.pack(pady=10)


        tk.Label(result_frame, text="Found ME Attributes:", font=("Arial", 12, "bold")).grid(row=0, column=0, sticky="w", padx=5)
        match_results_text = scrolledtext.ScrolledText(result_frame, wrap=tk.WORD, width=40, height=10, font=("Arial", 12))
        match_results_text.grid(row=1, column=0, padx=5, pady=5)

        tk.Label(result_frame, text="Not Found ME Attributes:", font=("Arial", 12, "bold")).grid(row=0, column=1, sticky="w", padx=5)
        missing_results_text = scrolledtext.ScrolledText(result_frame, wrap=tk.WORD, width=40, height=10, font=("Arial", 12))
        missing_results_text.grid(row=1, column=1, padx=5, pady=5)

    def search_text(self):
        keyword = self.search_var.get()
        self.search_results.clear()
        self.packet_display.tag_remove("highlight", "1.0", tk.END)
        if not keyword:
            return
        start = "1.0"
        while True:
            start = self.packet_display.search(keyword, start, stopindex=tk.END, nocase=True)
            if not start:
                break
            end = f"{start}+{len(keyword)}c"
            self.packet_display.tag_add("highlight", start, end)
            self.search_results.append((start, end))
            start = end
        self.packet_display.tag_config("highlight", background="yellow", foreground="black")
        if self.search_results:
            self.current_search_index = 0
            self.goto_search_result()
        self.update_search_count()

    def search_prev(self):
        if self.search_results and self.current_search_index > 0:
            self.current_search_index -= 1
            self.goto_search_result()
        self.update_search_count()

    def search_next(self):
        if self.search_results and self.current_search_index < len(self.search_results) - 1:
            self.current_search_index += 1
            self.goto_search_result()
        self.update_search_count()

    def goto_search_result(self):
        start, end = self.search_results[self.current_search_index]
        self.packet_display.tag_remove("current", "1.0", tk.END)
        self.packet_display.tag_add("current", start, end)
        self.packet_display.tag_config("current", background="orange", foreground="black")
        self.packet_display.mark_set(tk.INSERT, start)
        self.packet_display.see(start)

    def update_search_count(self):
        total = len(self.search_results)
        current = self.current_search_index + 1 if self.search_results else 0
        self.search_count_label.config(text=f"{current} / {total}")



    def open_release_note(self):
        release_note_content = """
        #########################################################
        # PON OMCI packets analysis tool                        #                    
        # Version : v1                                          #            
        # Date : 2024.9.9                                       #                                                                   
        # Releaser: Selina Wu                                   #                                 
        #########################################################
        ---------------------------------------------------------
        Version : v1.0                                                                               
        Date : 2024.9.9
        Purpose of the Release: 
        1.Main Functions added:
        [Menu] 
        Create the Menu of Save New File/Open File/Save as File
        Create the Action Menu of Undo/Redo/Cut/Copy/Paste 
        Add Testcases on TestRail Function
        Add Release_note

        2.OMCI Functions:
        Open Wireshark File to analyze
        Compare Files

        Releaser: Selina Wu
        ---------------------------------------------------------
        Version : v1.1                                                                               
        Date : 2024.11.15
        Purpose of the Release: 
        1.Remove unnecessary the  Undo/Redo/Cut/Copy/Paste Action Menu
        2.Modified Testcases on TestRail Function

        Releaser: Selina Wu
        ---------------------------------------------------------
        Version : v2.0                                                                               
        Date : 2024.12.18
        Purpose of the Release: 
        1.Create TestRail Function : select a test case from TestRail
        to verify the result PASS or FAIL

        a. Select a Test Case GUI: Drop-down form to select a test case 
        and press Test button

        b.Test Case Validation GUI: press the Validate button 
        to check Found or Not Found ME Attributes and show PASS or FAIL result

        Releaser: Selina Wu
        ---------------------------------------------------------
        Version : v3.0                                                                               
        Date : 2024.12.26
        Purpose of the Release: 
        Create Generate OMCI Analysis Table Function : 
        Generate a Excel table of OMCI data and save to the root directory.

        Releaser: Selina Wu
        ---------------------------------------------------------
        Version : v4.0                                                                               
        Date : 2024.12.31
        Purpose of the Release:  
        1.Change the Generate_omci_analysis_table to save the new file in the root directory
        2.Optimize Excel table content: enabled filter in first row, set default view to 80%,
        ME Class Upload Content column, and Mark the important parameters of the table content 
        with red letters on a yellow background

        Releaser: Selina Wu
        ---------------------------------------------------------

        """
          #  messagebox.showinfo("Release Note", release_note)

        self.temp_file_path = os.path.join(os.getenv('TEMP'), 'release_note.txt')

        # Write the release note content to the temporary file
        with open(self.temp_file_path, 'w') as file:
            file.write(release_note_content)

         # Open the release note in Notepad
        os.system(f'notepad.exe {self.temp_file_path}')


 

if __name__ == "__main__":
    root = tk.Tk()
    app = OMCIAnalyzerApp(root)
    root.mainloop()
